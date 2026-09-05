"""年度汇总页：全年 12 个月考勤与工资一览表，支持导出 Excel。"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from PySide6.QtCore import Qt
from PySide6.QtGui import QColor
from PySide6.QtWidgets import (
    QHBoxLayout, QHeaderView, QLabel, QPushButton, QFileDialog,
    QTableWidget, QTableWidgetItem, QSizePolicy, QMessageBox,
)

from . import calc
from .excel_style import (
    BORDER, BOLD_FONT, BODY_FONT, BAD_FONT, OK_FONT,
    HEADER_FILL, HEADER_FONT, MONEY_FMT, INT_FMT, FLT_FMT,
)
from .widgets import Card

_COLS = ["月份", "出勤(天)", "加班(小时)", "请假(小时)", "应发工资",
         "个人扣除", "请假扣款", "实发工资", "公司成本", "工时合规"]
_NUM_COLS = list(range(1, len(_COLS) - 1))  # 数值列索引
_COL_WIDTHS = [10, 10, 12, 12, 14, 14, 14, 14, 14, 12]  # 与 _COLS 一一对应


class AnnualPageMixin:
    def _annual_shift(self, delta):
        if self._book is None:
            return
        self._annual_year = int(self._annual_year or self._book.year) + delta
        self._fill_annual()

    def _annual_rows(self, year: int):
        """读取全年 12 个月并核算。返回 (行数据, 合计)；缺失月份各项为 None。"""
        rows = []
        totals = [0.0] * (len(_COLS) - 2)  # 出勤..公司成本
        for m in range(1, 13):
            book = self.store.load(year, m)
            if book is None:
                rows.append(None)
                continue
            r = calc.compute(book)
            vals = [float(r.counts.work), r.total_daily_ot_hours, r.daily_leave_hours,
                    r.gross_wage, r.personal_deductions_total, r.leave_deduction_total,
                    r.take_home, r.gross_wage + r.company_social + r.company_fund]
            rows.append((m, vals, r.work_time_legality))
            for i, v in enumerate(vals):
                totals[i] += v
        return rows, totals

    def _fill_annual(self):
        if self._book is None or not hasattr(self, "annual_lay"):
            return
        if not self._annual_year:
            self._annual_year = self._book.year
        lay = self.annual_lay
        self._clear_layout(lay)
        year = int(self._annual_year)

        # —— 顶栏：年份切换 + 打印 ——
        bar = QHBoxLayout()
        bar.setSpacing(8)
        prev_y = QPushButton("‹ 上一年")
        next_y = QPushButton("下一年 ›")
        prev_y.clicked.connect(lambda: self._annual_shift(-1))
        next_y.clicked.connect(lambda: self._annual_shift(1))
        title = QLabel(f"{year} 年 · 12 个月一览")
        title.setObjectName("cardTitle")
        printb = QPushButton("📊 导出 Excel")
        printb.setObjectName("ghost")
        printb.setCursor(Qt.PointingHandCursor)
        printb.clicked.connect(self._export_annual_xlsx)
        bar.addWidget(prev_y)
        bar.addWidget(title)
        bar.addWidget(next_y)
        bar.addStretch(1)
        bar.addWidget(printb)
        lay.addLayout(bar)

        # —— 表格卡片（Card 化）——
        card = Card(title="", variant="default", margins=(14, 12, 14, 12))
        cv = card.body_layout()
        cv.setSpacing(6)
        hint = QLabel("数据读取自已保存的月份；未保存的月份显示 “—”。公司成本 = 应发 + 公司社保 + 公司公积金。")
        hint.setObjectName("secHint")
        cv.addWidget(hint)

        rows, totals = self._annual_rows(year)
        self._annual_cache = (year, rows, totals)  # 缓存供打印复用
        table = QTableWidget(13, len(_COLS))
        table.setHorizontalHeaderLabels(_COLS)
        table.verticalHeader().setVisible(False)
        table.setEditTriggers(QTableWidget.NoEditTriggers)
        table.setSelectionMode(QTableWidget.NoSelection)
        table.setFocusPolicy(Qt.NoFocus)
        header = table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Stretch)
        # 允许表头在高 DPI 或不同布局中伸缩，但保留最小高度
        header.setMinimumHeight(30)
        header.setMaximumHeight(60)
        header.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

        for m in range(1, 13):
            row = m - 1
            data = rows[row]
            self._annual_cell(table, row, 0, f"{m} 月", center=True)
            if data is None:
                for c in _NUM_COLS:
                    self._annual_cell(table, row, c, "—", dim=True)
                self._annual_cell(table, row, len(_COLS) - 1, "—", dim=True, center=True)
                continue
            _, vals, legal = data
            texts = [f"{vals[0]:g}", f"{vals[1]:.1f}", f"{vals[2]:.1f}",
                     f"{vals[3]:,.2f}", f"{vals[4]:,.2f}", f"{vals[5]:,.2f}",
                     f"{vals[6]:,.2f}", f"{vals[7]:,.2f}"]
            for c, t in enumerate(texts, start=1):
                self._annual_cell(table, row, c, t)
            ok = "不违法" in legal
            self._annual_cell(table, row, len(_COLS) - 1, "合规" if ok else "违法",
                              color=("#0B7A3B" if ok else "#C00000"), bold=True, center=True)

        # 合计行
        last = 12
        self._annual_cell(table, last, 0, "合计", bold=True, center=True)
        for i, v in enumerate(totals, start=1):
            t = f"{v:g}" if i <= 3 else f"{v:,.2f}"
            self._annual_cell(table, last, i, t, bold=True)
        self._annual_cell(table, last, len(_COLS) - 1, "", dim=True)

        cv.addWidget(table, 1)
        lay.addWidget(card, 1)
        self._annual_table = table

    @staticmethod
    def _annual_cell(table, row, col, text, bold=False, dim=False,
                     color=None, center=False):
        it = QTableWidgetItem(text)
        it.setFlags(Qt.ItemIsEnabled)  # 只读
        f = it.font()
        f.setBold(bold)
        f.setPointSize(f.pointSize() - (1 if dim else 0))
        it.setFont(f)
        it.setForeground(QColor(Qt.gray) if dim else QColor(color or "#0F172A"))
        if center:
            it.setTextAlignment(Qt.AlignCenter)
        else:
            it.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
        table.setItem(row, col, it)

    def _export_annual_xlsx(self):
        """把全年汇总写到 .xlsx（替代原系统打印对话框；用户选保存位置）。"""
        if self._book is None:
            return
        year = int(self._annual_year or self._book.year)
        cache = getattr(self, "_annual_cache", None)
        if cache is not None and cache[0] == year:
            rows, totals = cache[1], cache[2]
        else:
            rows, totals = self._annual_rows(year)
        default_name = f"年度汇总_{year}年.xlsx"
        path, _ = QFileDialog.getSaveFileName(
            self, "导出年度汇总", default_name, "Excel 工作簿 (*.xlsx)")
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"
        try:
            _write_annual_xlsx(path, year, rows, totals)
        except Exception as ex:
            QMessageBox.warning(self, "导出失败", str(ex))
            return
        QMessageBox.information(self, "已导出", f"已写入 {path}")


def _write_annual_xlsx(path: str, year: int, rows, totals):
    """把年度汇总渲染到 .xlsx（独立函数，便于复用 / 测试）。"""
    wb = Workbook()
    ws = wb.active
    ws.title = f"{year}年汇总"

    # —— 表头 ——
    n_cols = len(_COLS)
    for c, name in enumerate(_COLS, start=1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "B2"  # 冻结首行 + 首列

    # —— 12 个月数据 ——

    def _put(row, col, value, *, kind, align="right"):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = BODY_FONT
        cell.border = BORDER
        if kind == "money":
            cell.number_format = MONEY_FMT
        elif kind == "int":
            cell.number_format = INT_FMT
        elif kind == "flt":
            cell.number_format = FLT_FMT
        cell.alignment = Alignment(horizontal=align, vertical="center")

    def _put_text(row, col, value, *, font=None, align="right"):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = font or BODY_FONT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal=align, vertical="center")

    for m in range(1, 13):
        r = m + 1  # 数据行：2..13（第 1 行是表头）
        data = rows[m - 1]
        # A 列：月份（居中）
        mc = ws.cell(row=r, column=1, value=f"{m} 月")
        mc.font = BODY_FONT
        mc.border = BORDER
        mc.alignment = Alignment(horizontal="center", vertical="center")

        if data is None:
            for c in _NUM_COLS:
                _put_text(r, c, "—")
            _put_text(r, n_cols, "—")
            continue

        _, vals, legal = data
        # B 出勤(天) int；C/D 加班/请假 小时 flt；E..I 金额
        _put(r, 2, vals[0], kind="int")
        _put(r, 3, vals[1], kind="flt")
        _put(r, 4, vals[2], kind="flt")
        _put(r, 5, vals[3], kind="money")
        _put(r, 6, vals[4], kind="money")
        _put(r, 7, vals[5], kind="money")
        _put(r, 8, vals[6], kind="money")
        _put(r, 9, vals[7], kind="money")
        # J 工时合规
        ok = "不违法" in legal
        text = "合规" if ok else "违法"
        _put_text(r, n_cols, text, font=OK_FONT if ok else BAD_FONT,
                  align="center")
        ws.row_dimensions[r].height = 20

    # —— 合计行（第 14 行）——
    total_row = 14
    tc = ws.cell(row=total_row, column=1, value="合计")
    tc.font = BOLD_FONT
    tc.border = BORDER
    tc.alignment = Alignment(horizontal="center", vertical="center")
    # 出勤(天) 整数；加班/请假 一位小数；其余 金额
    for i, v in enumerate(totals, start=2):
        # 按 _NUM_COLS 的索引判定格式：cols 1,2,3 对应 vals 的 出勤/加班/请假
        # totals 顺序与 vals 一致
        if i == 2:
            kind = "int"
        elif i in (3, 4):
            kind = "flt"
        else:
            kind = "money"
        cell = ws.cell(row=total_row, column=i, value=v)
        cell.font = BOLD_FONT
        cell.border = BORDER
        cell.number_format = MONEY_FMT if kind == "money" else (
            INT_FMT if kind == "int" else FLT_FMT)
        cell.alignment = Alignment(horizontal="right", vertical="center")
    # 工时合规 合计：留空
    ws.cell(row=total_row, column=n_cols, value="").border = BORDER
    ws.row_dimensions[total_row].height = 22

    # —— 列宽 + 表头底色覆盖（freeze 已设置）——
    for idx, w in enumerate(_COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    wb.save(path)
