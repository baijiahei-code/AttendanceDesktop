"""报表页渲染：折叠明细组 + 右侧汇总。"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment

from PySide6.QtCore import Qt
from PySide6.QtWidgets import (
    QFrame, QGridLayout, QHBoxLayout, QLabel, QPushButton,
    QScrollArea, QVBoxLayout, QWidget, QSizePolicy,
    QFileDialog, QMessageBox,
)

# 公共 Excel 样式（颜色/字体/边框/数字格式）—— 改色只需改 app/excel_style.py
from .excel_style import (
    BOLD_FONT, BODY_FONT, BAD_FONT, OK_FONT, INFO_FONT,
    TITLE_FONT, SECTION_FILL, GROUP_FILL,
    OK_FILL, BAD_FILL, MONEY_FMT, INT_FMT, FLT_FMT,
)
from .widgets import Card


class ReportPageMixin:
    def _fill_report(self):
        page = self.report_page
        if not hasattr(self, "_report_lay"):
            self._report_lay = QVBoxLayout(page)
            self._report_lay.setContentsMargins(0, 0, 0, 0)
            self._report_lay.setSpacing(0)
        lay = self._report_lay
        self._clear_layout(lay)
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        self.report_root = QWidget()
        scroll.setWidget(self.report_root)
        self.report_lay = QVBoxLayout(self.report_root)
        self.report_lay.setContentsMargins(20, 14, 20, 18)
        self.report_lay.setSpacing(10)
        lay.addWidget(scroll)

    def _render_report(self, r):
        if not hasattr(self, "report_lay"):
            return
        c = r.counts
        lay = self.report_lay
        self._clear_layout(lay)
        if not hasattr(self, "_group_bodies"):
            self._group_bodies = {}

        body = QHBoxLayout()
        body.setSpacing(12)

        # 左：明细分组（纵向，标题可折叠）
        left = QWidget()
        ll = QVBoxLayout(left)
        ll.setContentsMargins(0, 0, 0, 0)
        ll.setSpacing(10)
        for i, g in enumerate(r.groups):
            ll.addWidget(self._report_group_card(g, i))
        ll.addStretch(1)
        body.addWidget(left, 1)

        # 右：常驻汇总卡（Card 化）
        side = Card(title="", variant="default", margins=(16, 14, 16, 14))
        # 右侧汇总卡：允许伸缩显示，避免窄窗口时被裁切
        side.setMinimumWidth(260)
        side.setMaximumWidth(380)
        side.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Expanding)
        self._side_card(side, r)
        body.addWidget(side, 0)
        lay.addLayout(body, 1)

        # 考勤页统计条
        if hasattr(self, "cal_count_label"):
            self.cal_count_label.setText(
                f"上班 {c.work} · 休息 {c.rest} · 事假 {c.personal_leave} · 病假 {c.sick_leave} · "
                f"婚/丧/产/年假 {c.marriage_leave + c.bereavement_leave + c.maternity_leave + c.annual_leave} · "
                f"法定节假日 {c.legal_holiday} · 加班 {r.total_daily_ot_hours:.1f}h · 请假 {r.daily_leave_hours:.1f}h")

    def _report_group_card(self, g, idx):
        # Card 化 —— 报表分组卡（保留可点击标题实现折叠/展开）
        f = Card(title="", variant="default", margins=(14, 8, 14, 12))
        vl = f.body_layout()
        vl.setSpacing(6)
        head = QPushButton("▾ " + g["title"])
        head.setCursor(Qt.PointingHandCursor)
        head.setStyleSheet(
            "QPushButton{background:transparent;border:none;color:#0F172A;"
            "font-weight:700;font-size:13.5px;text-align:left;padding:2px 0;}")
        head.setObjectName("reportGroupHead")
        body_w = QWidget()
        bl = QGridLayout(body_w)
        bl.setContentsMargins(0, 0, 0, 0)
        bl.setHorizontalSpacing(16)
        bl.setVerticalSpacing(4)
        for row, line in enumerate(g["lines"]):
            lab = QLabel(line["label"])
            lab.setStyleSheet("color:#475569;")
            vl2 = QLabel(self._fmt_line(line))
            color, bold = self._line_style(line)
            vl2.setStyleSheet(f"color:{color};font-weight:{'700' if bold else '400'};")
            vl2.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            bl.addWidget(lab, row, 0)
            bl.addWidget(vl2, row, 1)
        self._group_bodies[idx] = body_w

        def toggle(_=False, i=idx):
            bw = self._group_bodies.get(i)
            if bw is None:
                return
            show = not bw.isVisible()
            bw.setVisible(show)
            head.setText(("▾ " if show else "▸ ") + g["title"])

        head.clicked.connect(toggle)
        vl.addWidget(head)
        vl.addWidget(body_w)
        return f

    def _side_card(self, side, r):
        c = r.counts
        # side 现在是 Card；拿到它的 body layout 用作容器
        sl = side.body_layout()
        sl.setSpacing(5)

        def small(txt, _=None):
            l = QLabel(txt)
            l.setStyleSheet("color:#667085;font-size:12px;")
            sl.addWidget(l)
            return l

        def row(txt, val, color):
            h = QHBoxLayout()
            a = QLabel(txt)
            a.setStyleSheet("color:#667085;")
            b = QLabel(val)
            b.setStyleSheet(f"color:{color};font-weight:700;")
            b.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            h.addWidget(a)
            h.addStretch(1)
            h.addWidget(b)
            sl.addLayout(h)

        def hr():
            line = QFrame()
            line.setFrameShape(QFrame.HLine)
            line.setStyleSheet("color:#E6E8F0;")
            sl.addWidget(line)

        small("应发工资")
        gv = QLabel(f"¥ {r.gross_wage:,.2f}")
        gv.setStyleSheet("color:#4F46E5;font-size:26px;font-weight:800;")
        sl.addWidget(gv)
        row("个人扣除", f"- ¥ {r.personal_deductions_total:,.2f}", "#F04438")
        row("请假扣款", f"- ¥ {r.leave_deduction_total:,.2f}", "#F04438")
        hr()
        small("实际到手")
        tv = QLabel(f"¥ {r.take_home:,.2f}")
        tv.setStyleSheet("color:#4F46E5;font-size:30px;font-weight:800;")
        sl.addWidget(tv)
        hr()
        row("到手小时工资", f"¥ {r.take_home_hourly:,.2f} / h", "#0F172A")
        row("提供正常劳动", f"{c.normal_labor_days} 天", "#0F172A")
        row("公司总成本", f"¥ {r.gross_wage + r.company_social + r.company_fund:,.2f}", "#0F172A")
        tip = QLabel("工作日加班 ×1.5 · 休息日 ×2 · 法定节假日 ×3 · "
                     "请假按（应发−个人扣除）÷ 约定工作天数计日扣款")
        tip.setWordWrap(True)
        tip.setStyleSheet("color:#98A2B3;font-size:11px;margin-top:6px;")
        sl.addWidget(tip)
        sl.addStretch(1)

        # —— 导出 Excel ——
        # 锁定月份也能导出 → 用 add_widget_untracked 让 Card 不追踪、不参与 set_locked
        pbtn = QPushButton("📊 导出 Excel")
        pbtn.setObjectName("primary")
        pbtn.setCursor(Qt.PointingHandCursor)
        pbtn.setStyleSheet("QPushButton#primary{min-height:30px; border-radius:9px;}")
        pbtn.clicked.connect(self._export_report_xlsx)
        side.add_widget_untracked(pbtn)

    @staticmethod
    def _fmt_line(line):
        if line["text"] is not None:
            return line["text"]
        v = line["value"]
        # 如果值为 None，不显示占位符（返回空字符串）
        if v is None:
            s = ""
        else:
            s = f"{v:,.2f}" if isinstance(v, float) else str(v)
            if line["unit"]:
                s += " " + line["unit"]
        return s

    @staticmethod
    def _line_style(line):
        if line["kind"] == "ok":
            return "#0B7A3B", True
        if line["kind"] == "bad":
            return "#C00000", True
        if line["bold"]:
            return "#1D4ED8", True
        return "#0F172A", False

    def _export_report_xlsx(self):
        """把当前月份报表写到 .xlsx，多 sheet：摘要 / 明细 / 合规。"""
        if self._last_result is None or self._book is None:
            return
        b = self._book
        r = self._last_result
        default_name = f"工资考勤报表_{b.year}年{b.month}月.xlsx"
        path, _ = QFileDialog.getSaveFileName(
            self, "导出报表", default_name, "Excel 工作簿 (*.xlsx)")
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"
        try:
            _write_report_xlsx(path, r, b)
        except Exception as ex:
            QMessageBox.warning(self, "导出失败", str(ex))
            return
        QMessageBox.information(self, "已导出", f"已写入 {path}")


def _write_report_xlsx(path: str, r, b):
    """把单月报表渲染到单个 sheet，纵向排版：标题/备注 → 汇总 → 明细分组 → 合规判定。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "工资考勤报表"

    NCOL = 3  # A=标签 B=数值/文本 C=单位

    def write_section_title(row, title):
        """区域标题：合并 A:C，浅靛紫底色 + 加粗。"""
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOL)
        c = ws.cell(row=row, column=1, value=title)
        c.font = BOLD_FONT
        c.fill = SECTION_FILL
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 22

    def write_kv(row, label, value, *, unit="", font=None, fmt=None, align_value="left"):
        """通用两列或三列行：A 标签 / B 数值 / C 单位。"""
        a = ws.cell(row=row, column=1, value=label)
        a.font = INFO_FONT
        a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        b_cell = ws.cell(row=row, column=2, value=value)
        b_cell.font = font or BODY_FONT
        b_cell.alignment = Alignment(horizontal=align_value, vertical="center", indent=1)
        if fmt:
            b_cell.number_format = fmt
        if unit:
            u = ws.cell(row=row, column=3, value=unit)
            u.font = INFO_FONT
            u.alignment = Alignment(horizontal="left", vertical="center")

    def blank(row):
        ws.row_dimensions[row].height = 8

    # ---- 标题 + 备注（合并 A:C） ----
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NCOL)
    t = ws.cell(row=1, column=1,
                value=f"工资考勤表 · {b.year} 年 {b.month} 月 核算报表")
    t.font = TITLE_FONT
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    note = (b.note or "").strip() or "（无）"
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NCOL)
    n = ws.cell(row=2, column=1, value=f"备注：{note}")
    n.font = INFO_FONT
    n.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=1)
    ws.row_dimensions[2].height = 22

    company_cost = r.gross_wage + r.company_social + r.company_fund
    cur = 4  # 当前写入行

    # ---- 段一：汇总（关键数字） ----
    write_section_title(cur, "📌 汇总数据"); cur += 1
    write_kv(cur, "应发工资",     r.gross_wage,
             font=BOLD_FONT, fmt=MONEY_FMT); cur += 1
    write_kv(cur, "个人扣除",     -r.personal_deductions_total,
             font=BOLD_FONT, fmt=MONEY_FMT); cur += 1
    write_kv(cur, "请假扣款",     -r.leave_deduction_total,
             font=BOLD_FONT, fmt=MONEY_FMT); cur += 1
    write_kv(cur, "实际到手",     r.take_home,
             font=BOLD_FONT, fmt=MONEY_FMT); cur += 1
    write_kv(cur, "到手小时工资", r.take_home_hourly,
             unit="元 / h", fmt=MONEY_FMT); cur += 1
    write_kv(cur, "公司总成本",   company_cost,
             unit="含公司社保 / 公积金", fmt=MONEY_FMT); cur += 1
    blank(cur); cur += 1

    # ---- 段二：明细分组 ----
    write_section_title(cur, "📋 明细分组"); cur += 1
    for g in r.groups:
        # 分组小标题：A:C 合并 + 加粗灰底
        ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=NCOL)
        gh = ws.cell(row=cur, column=1, value="  " + g["title"])
        gh.font = BOLD_FONT
        gh.fill = GROUP_FILL
        gh.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[cur].height = 20
        cur += 1
        for line in g["lines"]:
            label = line["label"]
            kind = line["kind"]
            bold = line["bold"]
            if line["text"] is not None:
                write_kv(cur, label, line["text"],
                         font=(OK_FONT if kind == "ok" else
                               BAD_FONT if kind == "bad" else
                               INFO_FONT if kind == "i" else
                               (BOLD_FONT if bold else BODY_FONT)),
                         align_value="left")
            else:
                v = line["value"]
                if v is None:
                    write_kv(cur, label, "")
                else:
                    if isinstance(v, float):
                        if line["unit"] == "天" and float(v).is_integer():
                            fmt = INT_FMT
                        elif line["unit"] == "小时":
                            fmt = FLT_FMT
                        else:
                            fmt = MONEY_FMT
                    elif isinstance(v, int):
                        fmt = INT_FMT
                    else:
                        fmt = None
                    write_kv(cur, label, v, unit=line["unit"] or "",
                             font=(BOLD_FONT if bold else BODY_FONT),
                             fmt=fmt, align_value="right")
            cur += 1
    blank(cur); cur += 1

    # ---- 段三：合规判定 ----
    write_section_title(cur, "⚖️ 合规判定"); cur += 1

    def emit(k, status, remark, kind):
        nonlocal cur
        # 标签 = 判定维度
        a = ws.cell(row=cur, column=1, value=k)
        a.font = BOLD_FONT
        a.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        # B 列：判定结果（合并 B:C 用于放长文本 remarks）
        ws.merge_cells(start_row=cur, start_column=2, end_row=cur, end_column=NCOL)
        s = ws.cell(row=cur, column=2, value=status)
        s.font = (OK_FONT if kind == "ok" else
                  BAD_FONT if kind == "bad" else INFO_FONT)
        s.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        if kind == "ok":
            s.fill = OK_FILL
        elif kind == "bad":
            s.fill = BAD_FILL
        # 备注写到下一行（B:C 合并）
        cur += 1
        if remark:
            ws.merge_cells(start_row=cur, start_column=2, end_row=cur, end_column=NCOL)
            rc = ws.cell(row=cur, column=2, value=remark)
            rc.font = INFO_FONT
            rc.alignment = Alignment(horizontal="left", vertical="center",
                                     wrap_text=True, indent=1)
            ws.row_dimensions[cur].height = max(20, 16 * ((len(remark) // 70) + 1))

    emit("最低工资标准",
         r.min_wage_status,
         ("计入部分 ≥ 地方最低月工资标准" if "不低于" in r.min_wage_status
          else "工资构成部分低于地方最低月工资标准"),
         "ok" if "不低于" in r.min_wage_status else "bad")
    cur += 1
    emit("工时与加班合规",
         r.work_time_legality,
         ("；".join(r.work_time_issues) if r.work_time_issues
          else "未触发工时违规条款"),
         "ok" if "不违法" in r.work_time_legality else "bad")
    cur += 1
    emit("公司雇佣成本（仅参考）",
         r.company_cost_status,
         "非全日制小时最低工资参考",
         "i")
    cur += 1
    legal_standard = ("月工时≤220h / 月加班≤36h / 上班≤26天 / 单日加班≤3h；"
                      "工作日加班×1.5 · 休息日×2.0 · 法定节假日×3.0")
    emit("判定标准", "—", legal_standard, "i")
    cur += 1

    # 列宽
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 28

    wb.save(path)
