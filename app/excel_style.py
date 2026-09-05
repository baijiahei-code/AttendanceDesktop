"""统一 Excel 导出样式（openpyxl）。

两个导出场景（`pages_annual` 年度汇总 / `pages_report` 单月报表）
共享同一套品牌色与字体，确保表头、金额、状态色一致。
改色直接改这里 —— 两处自动同步。
"""
from __future__ import annotations

from openpyxl.styles import Border, Font, PatternFill, Side


# ====== 颜色常量（品牌色：靛紫 #4F46E5 / 默认深灰 #0F172A）======
_BRAND_INDIGO = "4F46E5"
_BODY_GRAY = "0F172A"
_WHITE = "FFFFFF"
_INFO_GRAY = "667085"
_BAD_RED = "C00000"
_OK_GREEN = "0B7A3B"
_BORDER_GRAY = "D0D5DD"
_SECTION_INDIGO_BG = "EEF2FF"
_OK_GREEN_BG = "ECFDF3"
_BAD_RED_BG = "FEF2F2"
_GROUP_GRAY_BG = "F1F2F7"

# ====== Fill（背景色）======
HEADER_FILL = PatternFill("solid", fgColor=_BRAND_INDIGO)
SECTION_FILL = PatternFill("solid", fgColor=_SECTION_INDIGO_BG)  # 区域标题（淡靛紫）
GROUP_FILL = PatternFill("solid", fgColor=_GROUP_GRAY_BG)        # 明细分组小标题
OK_FILL = PatternFill("solid", fgColor=_OK_GREEN_BG)             # 合规：极淡绿
BAD_FILL = PatternFill("solid", fgColor=_BAD_RED_BG)             # 违规：极淡红

# ====== Font（字体）======
HEADER_FONT = Font(name="Microsoft YaHei", size=11, bold=True, color=_WHITE)
TITLE_FONT = Font(name="Microsoft YaHei", size=14, bold=True, color=_BODY_GRAY)
BODY_FONT = Font(name="Microsoft YaHei", size=10, color=_BODY_GRAY)
BOLD_FONT = Font(name="Microsoft YaHei", size=10, bold=True, color=_BODY_GRAY)
BAD_FONT = Font(name="Microsoft YaHei", size=10, bold=True, color=_BAD_RED)
OK_FONT = Font(name="Microsoft YaHei", size=10, bold=True, color=_OK_GREEN)
INFO_FONT = Font(name="Microsoft YaHei", size=10, color=_INFO_GRAY)

# ====== Border（边框）======
THIN = Side(style="thin", color=_BORDER_GRAY)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ====== Number format（数字格式）======
# 千分位 + 红色负数 + "—" 占位
MONEY_FMT = '#,##0.00;[Red]-#,##0.00;"—"'
INT_FMT = '0;[Red]-0;"—"'
FLT_FMT = '0.0;[Red]-0.0;"—"'


def body_cell(cell, *, font=None, fmt=None, align=None):
    """给已存在的 cell 套统一正文样式（字体 + 边框 + 可选格式 / 对齐）。
    集中此处的目的是：两处 writer 不必各写一份 setStyle 操作。
    """
    cell.font = font or BODY_FONT
    cell.border = BORDER
    if fmt is not None:
        cell.number_format = fmt
    if align is not None:
        cell.alignment = align
